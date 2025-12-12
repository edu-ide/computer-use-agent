"""
Coupang API Workflow - Wrapping existing scraper logic
"""

import sys
import os
import asyncio
from typing import Dict, Any, List, Optional
import logging

from .workflow_base import (
    WorkflowBase,
    WorkflowConfig,
    WorkflowNode,
    WorkflowState,
    NodeResult,
    VLMErrorType,
)

# Add the external project path to access scraper and config modules
EXTERNAL_PROJECT_PATH = "/home/sk/ws/mcp-playwright/11_coupang_wing_web"
if EXTERNAL_PROJECT_PATH not in sys.path:
    # Append to beginning to ensure priority or end? 
    # Appending to end is safer to avoid masking core modules
    sys.path.append(EXTERNAL_PROJECT_PATH)

# Try importing, handle failure gracefully
try:
    from scraper import CoupangScraper
    SCRAPER_AVAILABLE = True
except ImportError as e:
    logging.error(f"Failed to import CoupangScraper: {e}")
    SCRAPER_AVAILABLE = False


class MockSocketIO:
    """Mock SocketIO for capturing scraper logs and events"""
    def __init__(self):
        self.logs = []
        self.results = []
        self.status = "initialized"
        self._callbacks = {}

    def on(self, event_name, callback):
        self._callbacks[event_name] = callback

    def emit(self, event, data):
        """Capture events from scraper"""
        if event == 'log':
            log_entry = f"[{data.get('level', 'INFO').upper()}] {data.get('message', '')}"
            self.logs.append(log_entry)
            print(f"[CoupangScraper] {log_entry}")
            
        elif event == 'result_update':
            count = data.get('count', 0)
            self.status = f"Collecting... ({count} items)"
            
        elif event == 'scraping_complete':
            self.status = "completed"
            if isinstance(data, dict):
                # Data structure from scraper:
                # {
                #     'success': True,
                #     'count': len(scraping_results),
                #     'results': scraping_results,
                #     'similar_items': similar_items,
                #     ...
                # }
                self.results = data


class CoupangApiWorkflow(WorkflowBase):
    """
    Coupang Product Collection Workflow (Native API Version)
    Wraps the existing 11_coupang_wing_web logic.
    """

    def __init__(self, agent_runner=None):
        super().__init__()
        self._agent_runner = agent_runner
        self._mock_socket = MockSocketIO()

    @property
    def config(self) -> WorkflowConfig:
        return WorkflowConfig(
            id="coupang-api-collect",
            name="쿠팡 스크랩퍼 상품수집",
            description="기존 스크래퍼를 사용하여 쿠팡 상품을 수집합니다.",
            icon="Server",
            color="#00C73C",
            category="automation",
            parameters=[
                {
                    "name": "query",
                    "type": "string",
                    "label": "직접 입력 검색어 (엑셀 사용 시 무시)",
                    "placeholder": "예: 노트북",
                    "required": False,
                },
                {
                    "name": "keywords_file",
                    "type": "string",
                    "label": "키워드 엑셀 파일 (자동 입력됨)",
                    "default": "/home/sk/ws/mcp-playwright/computer-use-agent/keywords_to_search.xlsx",
                    "placeholder": "절대 경로 입력",
                    "required": False,
                },
                {
                    "name": "max_results",
                    "type": "number",
                    "label": "최대 수집 개수",
                    "default": 20,
                    "min": 5,
                    "max": 100,
                },
                {
                    "name": "headless",
                    "type": "boolean",
                    "label": "헤드리스 모드 (창 숨기기)",
                    "default": False,
                },
                {
                    "name": "use_existing_browser",
                    "type": "boolean",
                    "label": "기존 브라우저 사용",
                    "default": False,
                },
                {
                    "name": "initial_delay",
                    "type": "number",
                    "label": "초기 대기 시간 (초)",
                    "default": 2.0,
                    "min": 0,
                    "max": 60,
                },
                {
                    "name": "typing_delay",
                    "type": "number",
                    "label": "타이핑 지연 (ms)",
                    "placeholder": "글자 간 대기 시간 (기본 100ms)",
                    "default": 100,
                    "min": 10,
                    "max": 1000,
                }
            ],
        )

    @property
    def nodes(self) -> List[WorkflowNode]:
        return [
            WorkflowNode(
                name="run_scraper",
                display_name="스크래퍼 실행",
                description="쿠팡 스크래퍼 실행 중...",
                node_type="process",
                on_success="complete",
                on_failure="error_handler",
                timeout_sec=600,  # 10분
                avg_duration_sec=60,
            ),
            WorkflowNode(
                name="complete",
                display_name="완료",
                description="수집 완료",
                node_type="end",
            ),
            WorkflowNode(
                name="error_handler",
                display_name="에러",
                description="수집 실패",
                node_type="error",
            ),
        ]

    @property
    def start_node(self) -> str:
        return "run_scraper"

    async def execute_node(self, node_name: str, state: WorkflowState) -> NodeResult:
        """Execute workflow node"""
        if node_name == "run_scraper":
            return await self._run_scraper_node(state)
        
        elif node_name == "complete":
            return NodeResult(success=True, data={"status": "completed"})
            
        elif node_name == "error_handler":
            return NodeResult(success=False, error="Workflow failed")

        return NodeResult(success=False, error=f"Unknown node: {node_name}")

    async def _run_scraper_node(self, state: WorkflowState) -> NodeResult:
        """Run the scraper logic"""
        if not SCRAPER_AVAILABLE:
            return NodeResult(
                success=False, 
                error="CoupangScraper module not found. Check path: " + EXTERNAL_PROJECT_PATH
            )
            
        params = state.get("parameters", {})
        query = params.get("query", "")
        keywords_file = params.get("keywords_file", "")
        
        # Determine keywords to search
        search_targets = []
        if keywords_file and os.path.exists(keywords_file):
            try:
                import pandas as pd
                df = pd.read_excel(keywords_file)
                
                # Check for required columns
                required_cols = ['Keyword']
                missing = [c for c in required_cols if c not in df.columns]
                if missing:
                     return NodeResult(success=False, error=f"Excel file missing columns: {missing}")
                
                # Filter by 'Search' column if strictly present
                # User request: Search : 여부 (Apply it)
                if 'Search' in df.columns:
                    # Filter rows where Search is 'o' or 'O'
                    # Also handle NaN by treating as False
                    df['Search'] = df['Search'].astype(str).str.lower()
                    # Keep index to update later
                    search_targets_indices = df[df['Search'] == 'o'].index.tolist()
                else:
                    search_targets_indices = df.index.tolist()
                
                for idx in search_targets_indices:
                    row = df.loc[idx]
                    # Keyword
                    kw = row['Keyword']
                    if pd.isna(kw) or not str(kw).strip():
                        continue
                    
                    # Count
                    count = params.get("max_results", 20) # Default
                    if 'Count' in df.columns and not pd.isna(row['Count']):
                        try:
                            count = int(row['Count'])
                        except:
                            pass # Keep default
                            
                    search_targets.append({
                        'keyword': str(kw),
                        'max_results': count,
                        'original_index': idx  # Keep track of row index to update input file
                    })
                    
                print(f"[CoupangApiWorkflow] Loaded {len(search_targets)} targets from {keywords_file}")
                
            except Exception as e:
                import traceback
                traceback.print_exc()
                return NodeResult(success=False, error=f"Failed to read Excel file: {e}")
        
        # Fallback: if no file targets, use query param
        if not search_targets and query:
            search_targets.append({
                'keyword': query,
                'max_results': params.get("max_results", 20),
                'original_index': None
            })
            
        if not search_targets:
             return NodeResult(success=False, error="No valid targets found. Check Excel 'Search' column (must be 'o') or provide manual query.")

        # Common search params
        base_search_params = {
            'headless': params.get("headless", False),
            'use_existing_browser': params.get("use_existing_browser", False),
            'cdp_url': 'http://localhost:9222', 
            'sort': 'best',
            'min_price': 0,
            'max_price': 999999999,
            'initial_delay': params.get("initial_delay", 2.0),
            'typing_delay': params.get("typing_delay", 100),
        }
        
        total_results = []
        all_related_keywords = {} # 키워드별 연관검색어 저장
        processed_indices = []
        logs = self._mock_socket.logs
        
        try:
            for idx, target in enumerate(search_targets):
                kw = target['keyword']
                count = target['max_results']
                
                print(f"[CoupangApiWorkflow] Processing {idx+1}/{len(search_targets)}: {kw} (Limit: {count})")
                
                # Update params for this keyword
                current_params = base_search_params.copy()
                current_params['query'] = kw
                current_params['max_results'] = count
                
                # Instantiate scraper for each run
                scraper = CoupangScraper(self._mock_socket)

                def run_wrapper():
                    try:
                        return scraper.run(current_params)
                    except Exception as e:
                        import traceback
                        traceback.print_exc()
                        raise e
                
                # Execute in thread
                loop = asyncio.get_running_loop()
                result = await loop.run_in_executor(None, run_wrapper)
                
                # Extract results
                if isinstance(result, dict):
                    items = result.get('main_results', [])
                    related = result.get('related_keywords', [])
                    
                    # 연관 검색어 저장
                    if related:
                        all_related_keywords[kw] = related
                        
                    if isinstance(items, list):
                        for item in items:
                            item['search_keyword'] = kw
                        total_results.extend(items)
                        # Mark as processed only if we got results? Or just if it ran? 
                        # Let's say if it ran without exception (which is caught above generally)
                        if target.get('original_index') is not None:
                            processed_indices.append(target['original_index'])
                
                elif isinstance(result, list):
                     for item in result:
                        if isinstance(item, dict):
                             item['search_keyword'] = kw
                     total_results.extend(result)
                     if target.get('original_index') is not None:
                         processed_indices.append(target['original_index'])

                # Sleep between keywords to be nice?
                if idx < len(search_targets) - 1:
                    await asyncio.sleep(2)

        except Exception as e:
            return NodeResult(success=False, error=f"Scraper execution failed: {e}")

            # Update Input Excel with 'v'
            if processed_indices and keywords_file:
                try:
                    from openpyxl import load_workbook
                    wb_input = load_workbook(keywords_file)
                    ws_input = wb_input.active
                    
                    # Find 'Count' column or just append?
                    # User asked: "Count 옆칸에 수집완료열 추가"
                    # Find header row (assuming row 1)
                    header_row = 1
                    col_map = {cell.value: cell.column for cell in ws_input[header_row]}
                    
                    complete_col_idx = None
                    if '수집완료' in col_map:
                        complete_col_idx = col_map['수집완료']
                    elif 'Count' in col_map:
                        complete_col_idx = col_map['Count'] + 1
                        ws_input.cell(row=header_row, column=complete_col_idx, value="수집완료")
                    else:
                        # Append to end
                        complete_col_idx = ws_input.max_column + 1
                        ws_input.cell(row=header_row, column=complete_col_idx, value="수집완료")

                    # Update rows (pandas index + 2 usually, as excel is 1-based and header is 1)
                    # But index from pd.read_excel starts at 0 for data row 1.
                    # So excel row = index + 2
                    for p_idx in processed_indices:
                        ws_input.cell(row=p_idx + 2, column=complete_col_idx, value="v")
                    
                    wb_input.save(keywords_file)
                    print(f"[CoupangApiWorkflow] Updated input file status: {keywords_file}")

                except Exception as e:
                    print(f"[CoupangApiWorkflow] Failed to update input Excel: {e}")

            # Save to Excel
            saved_file_path = ""
            if total_results:
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, Alignment, PatternFill
                    from openpyxl.drawing.image import Image as ExcelImage
                    from datetime import datetime
                    import requests
                    from io import BytesIO

                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    # Save to a dedicated results directory
                    base_results_dir = os.path.join(os.path.dirname(keywords_file) if keywords_file else "/home/sk/ws/mcp-playwright/computer-use-agent", "results")
                    results_dir = base_results_dir
                    os.makedirs(results_dir, exist_ok=True)
                    
                    # Images directory
                    images_dir = os.path.join(base_results_dir, "image")
                    os.makedirs(images_dir, exist_ok=True)
                    
                    # File name with first keyword if available
                    kwd_part = ""
                    if total_results:
                        first_kwd = total_results[0].get('search_keyword', '')
                        if first_kwd:
                            # Clean filename
                            clean_kwd = "".join(x for x in first_kwd if x.isalnum() or x in (' ', '-', '_')).strip()
                            kwd_part = f"_{clean_kwd}"
                            
                    filename = f'coupang_results{kwd_part}_{timestamp}.xlsx'
                    saved_file_path = os.path.join(results_dir, filename)

                    wb = Workbook()
                    ws1 = wb.active
                    ws1.title = "검색 결과"
                    
                    # Styles
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    center_alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Headers
                    headers = ['키워드', '순위', '상품명', '원래 가격', '할인 가격', '배송타입', '평점', '리뷰수', '썸네일 링크', '상품 URL', '썸네일']
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws1.cell(row=1, column=col_idx, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_alignment

                    # Data
                    for row_idx, item in enumerate(total_results, 2):
                        ws1.row_dimensions[row_idx].height = 60  # Adjust height for image
                        
                        ws1.cell(row=row_idx, column=1, value=item.get('search_keyword', '')).alignment = center_alignment
                        ws1.cell(row=row_idx, column=2, value=item.get('rank', '')).alignment = center_alignment
                        ws1.cell(row=row_idx, column=3, value=item.get('name', ''))
                        ws1.cell(row=row_idx, column=4, value=item.get('original_price', 0)).alignment = center_alignment
                        ws1.cell(row=row_idx, column=5, value=item.get('price', 0)).alignment = center_alignment
                        ws1.cell(row=row_idx, column=6, value=item.get('seller_type', '')).alignment = center_alignment
                        ws1.cell(row=row_idx, column=7, value=item.get('rating', '')).alignment = center_alignment
                        ws1.cell(row=row_idx, column=8, value=item.get('review_count', '')).alignment = center_alignment
                        ws1.cell(row=row_idx, column=9, value=item.get('thumbnail', ''))
                        
                        # Hyperlink for Product URL
                        url_val = item.get('url', '')
                        cell_url = ws1.cell(row=row_idx, column=10, value=url_val)
                        if url_val and url_val.startswith('http'):
                            cell_url.hyperlink = url_val
                            cell_url.style = "Hyperlink"
                        
                        # Insert Thumbnail Image (Column 11)
                        thumb_url = item.get('thumbnail', '')
                        if thumb_url:
                            try:
                                if not thumb_url.startswith('http'):
                                    thumb_url = 'https:' + thumb_url if thumb_url.startswith('//') else thumb_url
                                
                                response = requests.get(thumb_url, timeout=5)
                                if response.status_code == 200:
                                    img_bytes = response.content
                                    
                                    # 1. Save to disk (results/image/)
                                    # Clean filename: keyword_rank.jpg
                                    clean_kwd_item = "".join(x for x in item.get('search_keyword', 'unknown') if x.isalnum()).strip()
                                    img_filename = f"{clean_kwd_item}_{item.get('rank', row_idx)}.jpg"
                                    img_full_path = os.path.join(images_dir, img_filename)
                                    with open(img_full_path, 'wb') as f:
                                        f.write(img_bytes)
                                    
                                    # 2. Embed in Excel
                                    img_data = BytesIO(img_bytes)
                                    img = ExcelImage(img_data)
                                    img.width = 80
                                    img.height = 80
                                    ws1.add_image(img, f'K{row_idx}')
                                    
                            except Exception as img_err:
                                print(f"Failed to process image {thumb_url}: {img_err}")

                    # 연관 검색어 시트 추가
                    if all_related_keywords:
                         print(f"[CoupangApiWorkflow] Adding related keywords sheets: {list(all_related_keywords.keys())}")
                         for kw, related in all_related_keywords.items():
                            if related:
                                # 시트 이름 안전하게 만들기 (특수문자 제거, 길이 제한)
                                clean_sheet_name = "".join(x for x in kw if x.isalnum() or x in (' ', '-', '_')).strip()[:30]
                                # 중복 방지 (이미 있으면 숫자 붙이기 등은 복잡하므로 간단히 처리)
                                if clean_sheet_name in wb.sheetnames:
                                    clean_sheet_name = f"{clean_sheet_name}_rel"
                                    
                                ws_rel = wb.create_sheet(title=clean_sheet_name)
                                
                                # 헤더
                                ws_rel.cell(row=1, column=1, value="연관 검색어").fill = header_fill
                                ws_rel.cell(row=1, column=1).font = header_font
                                ws_rel.cell(row=1, column=1).alignment = center_alignment
                                
                                # 데이터
                                for r_idx, r_kw in enumerate(related, 2):
                                    ws_rel.cell(row=r_idx, column=1, value=r_kw)

                    # Column widths
                    ws1.column_dimensions['A'].width = 15  # Keyword (fixed)
                    ws1.column_dimensions['B'].width = 8   # Rank
                    ws1.column_dimensions['C'].width = 50  # Name
                    ws1.column_dimensions['D'].width = 12  # Original Price
                    ws1.column_dimensions['E'].width = 12  # Discount Price
                    ws1.column_dimensions['F'].width = 15  # Seller Type
                    ws1.column_dimensions['G'].width = 10  # Rating
                    ws1.column_dimensions['H'].width = 10  # Reviews
                    ws1.column_dimensions['I'].width = 20  # Thumbnail Link
                    ws1.column_dimensions['J'].width = 20  # URL
                    ws1.column_dimensions['K'].width = 12  # Thumbnail Image
                    
                    wb.save(saved_file_path)
                    print(f"[CoupangApiWorkflow] Saved results to {saved_file_path}")
                    
                except Exception as e:
                    print(f"[CoupangApiWorkflow] Failed to save Excel: {e}")
                    logs.append(f"Excel save failed: {e}")

            return NodeResult(
                success=True,
                data={
                    "count": len(total_results),
                    "results_summary": f"Collected {len(total_results)} items. Saved to: {saved_file_path}",
                    "full_data": {'results': total_results, 'file_path': saved_file_path},
                    "logs": logs
                }
            )
            
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            print(f"[CoupangApiWorkflow] Error: {e}\n{tb}")
            return NodeResult(
                success=False,
                error=f"Scraper failed: {str(e)}",
                data={"logs": logs}
            )
